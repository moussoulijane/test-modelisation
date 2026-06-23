"""
Robust time-series forecasting pipeline.

This file keeps the original project shape:
- load the first Excel workbook from in/
- analyze all numeric indicators
- export one Excel report and Plotly HTML charts into out/

Main corrections versus the legacy version:
- no interpolation of closed calendar days
- no ML leakage: every model is trained on train only, then tested on a true holdout
- model selection uses forecast errors, not R2
- selected lags are estimated from the train window only
- peak-aware metrics and calendar lift are used for spiky series
- optional XGBoost and KNN analog-path candidates are included
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
import json
import math
import re
import warnings

import numpy as np
import pandas as pd
from scipy import stats
from scipy.signal import find_peaks
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.stattools import adfuller

from sklearn.ensemble import GradientBoostingRegressor, RandomForestRegressor
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

import plotly.graph_objects as go
from plotly.subplots import make_subplots

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


warnings.filterwarnings("ignore")
np.random.seed(42)


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "in"
OUTPUT_DIR = BASE_DIR / "out"


@dataclass
class PipelineConfig:
    """Centralized knobs for the forecasting run."""

    horizon: int = 30
    holdout_min: int = 30
    holdout_max: int = 90
    holdout_ratio: float = 0.2
    min_train_size: int = 180
    max_lag: int = 120
    max_selected_lags: int = 24
    peak_quantile: float = 0.90
    peak_mad_multiplier: float = 3.5
    peak_weight: float = 4.0
    complex_model_min_gain: float = 0.05
    analog_lookback: int = 30
    analog_neighbors: int = 12
    use_xgboost: bool = True
    random_state: int = 42


@dataclass
class CandidateResult:
    name: str
    model: Any
    metrics: Dict[str, float]
    predictions: pd.Series
    model_pack: Dict[str, Any]


class KNNAnalogPathModel:
    """
    KNN path analog model.

    It searches historical windows with a shape similar to the current recent
    path and averages their following paths. This is useful when peaks are
    recurring by pattern but hard to capture with a smooth parametric model.
    """

    def __init__(self, lookback: int = 30, neighbors: int = 12) -> None:
        self.lookback = int(lookback)
        self.neighbors = int(neighbors)
        self.vectors: Optional[np.ndarray] = None
        self.paths: Optional[np.ndarray] = None
        self.path_horizon: int = 0
        self.fitted: bool = False

    @staticmethod
    def _normalize_path(values: np.ndarray) -> np.ndarray:
        base = values[0] if abs(values[0]) > 1e-12 else 1.0
        rel = values / base - 1.0
        return rel - np.nanmean(rel)

    def fit(self, series: pd.Series, horizon: int) -> "KNNAnalogPathModel":
        clean = series.dropna().astype(float)
        values = clean.values
        horizon = int(horizon)
        if len(values) < self.lookback + horizon + 5:
            raise ValueError("Not enough history for KNN analog path")

        vectors: List[np.ndarray] = []
        paths: List[np.ndarray] = []
        for end in range(self.lookback, len(values) - horizon):
            window = values[end - self.lookback : end]
            future = values[end : end + horizon]
            if not np.all(np.isfinite(window)) or not np.all(np.isfinite(future)):
                continue
            base = window[-1] if abs(window[-1]) > 1e-12 else 1.0
            vectors.append(self._normalize_path(window))
            paths.append(future / base)

        if len(vectors) < 5:
            raise ValueError("Not enough analog windows")

        self.vectors = np.vstack(vectors)
        self.paths = np.vstack(paths)
        self.path_horizon = horizon
        self.fitted = True
        return self

    def forecast(self, history: pd.Series, horizon: int) -> np.ndarray:
        if not self.fitted or self.vectors is None or self.paths is None:
            raise ValueError("KNN analog model is not fitted")

        hist = history.dropna().astype(float).values
        if len(hist) < self.lookback:
            raise ValueError("Not enough history to forecast with analog path")

        current = self._normalize_path(hist[-self.lookback :])
        distances = np.sqrt(np.mean((self.vectors - current) ** 2, axis=1))
        k = max(1, min(self.neighbors, len(distances)))
        idx = np.argsort(distances)[:k]
        weights = 1.0 / (distances[idx] + 1e-8)
        weights = weights / weights.sum()
        rel_path = np.average(self.paths[idx], axis=0, weights=weights)

        base = hist[-1]
        path = base * rel_path
        if horizon <= len(path):
            return path[:horizon]

        extra = np.repeat(path[-1], horizon - len(path))
        return np.concatenate([path, extra])


def _safe_float(value: Any, default: float = np.nan) -> float:
    try:
        out = float(value)
        return out if np.isfinite(out) else default
    except Exception:
        return default


def normalize_column_name(name: Any) -> str:
    text = str(name).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def morocco_fixed_holidays(years: Iterable[int]) -> set[pd.Timestamp]:
    """Civil Moroccan bank holidays. Religious holidays can be added via CSV."""

    holidays: set[pd.Timestamp] = set()
    fixed_mmdd = [
        (1, 1),   # New year
        (1, 11),  # Independence manifesto
        (5, 1),   # Labour day
        (7, 30),  # Throne day
        (8, 14),  # Oued Ed-Dahab
        (8, 20),  # Revolution of the King and the People
        (8, 21),  # Youth day
        (11, 6),  # Green March
        (11, 18), # Independence day
    ]
    for year in years:
        for month, day in fixed_mmdd:
            holidays.add(pd.Timestamp(year=year, month=month, day=day).normalize())
    return holidays


def load_extra_holidays(base_dir: Path) -> set[pd.Timestamp]:
    """
    Optional extension point: put in/holidays_ma.csv with one date per row.
    This avoids hard-coding lunar holidays in the model code.
    """

    path = base_dir / "in" / "holidays_ma.csv"
    if not path.exists():
        return set()
    raw = pd.read_csv(path, header=None)
    dates = pd.to_datetime(raw.iloc[:, 0], errors="coerce").dropna()
    return {pd.Timestamp(d).normalize() for d in dates}


def make_future_business_dates(last_date: pd.Timestamp, horizon: int, base_dir: Path) -> pd.DatetimeIndex:
    start_year = int(last_date.year)
    years = range(start_year, start_year + 3)
    holidays = morocco_fixed_holidays(years) | load_extra_holidays(base_dir)
    dates: List[pd.Timestamp] = []
    current = pd.Timestamp(last_date).normalize()
    while len(dates) < horizon:
        current = current + pd.Timedelta(days=1)
        if current.weekday() >= 5:
            continue
        if current in holidays:
            continue
        dates.append(current)
    return pd.DatetimeIndex(dates)


class TimeSeriesAnalyzer:
    """Analyzer, model selector and forecaster for all numeric series."""

    def __init__(
        self,
        df: pd.DataFrame,
        date_col: Optional[str] = None,
        indicator_cols: Optional[Sequence[str]] = None,
        config: Optional[PipelineConfig] = None,
    ) -> None:
        self.config = config or PipelineConfig()
        self.df = df.copy()
        self.date_col = date_col
        self.indicator_cols = list(indicator_cols or [c for c in df.columns if c != date_col])
        self.horizon = self.config.horizon

        if date_col:
            self.df[date_col] = pd.to_datetime(self.df[date_col], errors="coerce")
            self.df = self.df.dropna(subset=[date_col])
            self.df = self.df.sort_values(date_col)
            self.df = self.df.groupby(date_col, as_index=False).mean(numeric_only=True)
            self.df = self.df.set_index(date_col)

        self.df = self.df.sort_index()
        self.results: Dict[str, Any] = {}
        self.models: Dict[str, Dict[str, Any]] = {}
        self.forecasts: Dict[str, pd.Series] = {}
        self.classifications: Dict[str, Dict[str, Any]] = {}
        self.holdout_predictions: Dict[str, pd.Series] = {}
        self.model_comparisons: Dict[str, List[Dict[str, Any]]] = {}
        self.failures: List[Dict[str, str]] = []

    def analyze_stationarity(self, series: pd.Series) -> Dict[str, Any]:
        clean = series.dropna().astype(float)
        if len(clean) < 25 or clean.nunique() < 3:
            return {
                "adf_statistic": np.nan,
                "p_value": np.nan,
                "is_stationary": False,
                "reason": "insufficient_data",
            }
        try:
            result = adfuller(clean)
            return {
                "adf_statistic": result[0],
                "p_value": result[1],
                "is_stationary": bool(result[1] < 0.05),
                "critical_values": result[4],
            }
        except Exception as exc:
            return {"adf_statistic": np.nan, "p_value": np.nan, "is_stationary": False, "reason": str(exc)}

    def detect_trend(self, series: pd.Series) -> Dict[str, Any]:
        clean = series.dropna().astype(float)
        window = clean.tail(min(180, len(clean)))
        if len(window) < 20:
            return {"trend_type": "insufficient_data", "slope": 0.0, "r2": 0.0, "p_value": np.nan}

        x = np.arange(len(window), dtype=float)
        y = window.values
        slope, intercept, r_value, p_value, _ = stats.linregress(x, y)
        denom = abs(np.nanmedian(y)) if abs(np.nanmedian(y)) > 1e-12 else 1.0
        normalized_slope = slope / denom
        if p_value > 0.05 or abs(normalized_slope) < 0.0002:
            trend_type = "no_trend"
        elif slope > 0:
            trend_type = "increasing"
        else:
            trend_type = "decreasing"
        return {
            "trend_type": trend_type,
            "slope": slope,
            "normalized_slope": normalized_slope,
            "intercept": intercept,
            "r2": r_value**2,
            "p_value": p_value,
        }

    def detect_seasonality(self, series: pd.Series) -> Dict[str, Any]:
        clean = series.dropna().astype(float)
        if len(clean) < 60 or not isinstance(clean.index, pd.DatetimeIndex):
            return {"has_seasonality": False, "strength": 0.0, "period": None}

        total_var = float(np.nanvar(clean.values))
        if total_var <= 1e-12:
            return {"has_seasonality": False, "strength": 0.0, "period": None}

        dow_means = clean.groupby(clean.index.dayofweek).mean()
        dom_bucket = pd.Series(np.where(clean.index.day >= 25, "eom", "normal"), index=clean.index)
        eom_means = clean.groupby(dom_bucket).mean()
        dow_strength = float(np.nanvar(dow_means.values) / total_var)
        eom_strength = float(np.nanvar(eom_means.values) / total_var)
        strength = max(dow_strength, eom_strength)
        period = 5 if dow_strength >= eom_strength else 20
        return {"has_seasonality": strength > 0.03, "strength": strength, "period": period}

    def detect_volatility(self, series: pd.Series) -> Dict[str, Any]:
        clean = series.dropna().astype(float)
        if len(clean) < 3:
            return {"volatility": 0.0, "coefficient_variation": 0.0, "volatility_type": "stable"}
        mean_abs = abs(float(clean.mean()))
        cv = float(clean.std() / mean_abs) if mean_abs > 1e-12 else np.inf
        if cv < 0.05:
            vol_type = "low"
        elif cv < 0.20:
            vol_type = "medium"
        else:
            vol_type = "high"
        return {
            "volatility": float(clean.std()),
            "coefficient_variation": cv,
            "volatility_type": vol_type,
            "max_drawdown": float((clean / clean.cummax() - 1.0).min()),
        }

    def detect_peaks(self, series: pd.Series) -> Dict[str, Any]:
        clean = series.dropna().astype(float)
        if len(clean) < 20:
            return {"threshold": np.nan, "peak_count": 0, "peak_ratio": 0.0, "peak_dates": []}

        threshold = self._peak_threshold(clean)
        peaks_idx, _ = find_peaks(clean.values, height=threshold)
        peak_dates = [clean.index[i] for i in peaks_idx]
        return {
            "threshold": threshold,
            "peak_count": int(len(peaks_idx)),
            "peak_ratio": float(len(peaks_idx) / len(clean)),
            "peak_dates": peak_dates,
        }

    def classify_series(self, series: pd.Series, indicator_name: str) -> Dict[str, Any]:
        stationarity = self.analyze_stationarity(series)
        trend = self.detect_trend(series)
        seasonality = self.detect_seasonality(series)
        volatility = self.detect_volatility(series)
        peaks = self.detect_peaks(series)
        zero_ratio = float((series.dropna() == 0).mean()) if len(series.dropna()) else 0.0

        if zero_ratio > 0.15:
            classification = "INTERMITTENT"
        elif peaks["peak_ratio"] > 0.04 or volatility["coefficient_variation"] > 0.25:
            classification = "PEAKY"
        elif trend["trend_type"] != "no_trend" and seasonality["has_seasonality"]:
            classification = "TREND_SEASONAL"
        elif trend["trend_type"] != "no_trend":
            classification = "TRENDING"
        elif seasonality["has_seasonality"]:
            classification = "SEASONAL"
        else:
            classification = "STABLE"

        return {
            "indicator": indicator_name,
            "classification": classification,
            "stationarity": stationarity,
            "trend": trend,
            "seasonality": seasonality,
            "volatility": volatility,
            "peaks": peaks,
            "zero_ratio": zero_ratio,
        }

    def _record_failure(self, indicator: str, model_name: str, reason: str) -> None:
        self.failures.append({"indicator": indicator, "model": model_name, "reason": reason[:300]})

    def _peak_threshold(self, series: pd.Series) -> float:
        clean = series.dropna().astype(float)
        if clean.empty:
            return np.nan
        median = float(clean.median())
        mad = float(np.median(np.abs(clean.values - median)))
        robust = median + self.config.peak_mad_multiplier * 1.4826 * mad
        quantile = float(clean.quantile(self.config.peak_quantile))
        return max(robust, quantile)

    def select_lags(self, series: pd.Series) -> List[int]:
        clean = series.dropna().astype(float)
        max_lag = min(self.config.max_lag, max(2, len(clean) // 3))
        mandatory = [1, 2, 3, 5, 10, 20, 21, 22, 30, 60]
        scores: List[Tuple[int, float]] = []

        for lag in range(1, max_lag + 1):
            aligned = pd.concat([clean, clean.shift(lag)], axis=1).dropna()
            if len(aligned) < 20 or aligned.iloc[:, 1].nunique() < 2:
                corr = 0.0
            else:
                corr = abs(float(aligned.iloc[:, 0].corr(aligned.iloc[:, 1])))
                corr = corr if np.isfinite(corr) else 0.0
            scores.append((lag, corr))

        selected = [lag for lag, _ in sorted(scores, key=lambda item: item[1], reverse=True)[: self.config.max_selected_lags]]
        selected.extend([lag for lag in mandatory if lag <= max_lag])
        selected = sorted(set(selected))
        return selected[: self.config.max_selected_lags]

    @staticmethod
    def _calendar_features_for_date(date_value: Any) -> Dict[str, float]:
        if not isinstance(date_value, pd.Timestamp):
            return {}
        day = int(date_value.day)
        month = int(date_value.month)
        dow = int(date_value.dayofweek)
        quarter_end_month = month in (3, 6, 9, 12)
        features: Dict[str, float] = {
            "dow": float(dow),
            "day": float(day),
            "month": float(month),
            "is_month_end_window": float(day >= 25),
            "is_month_start": float(day <= 5),
            "is_quarter_end_window": float(quarter_end_month and day >= 20),
            "is_august": float(month == 8),
            "month_sin": math.sin(2 * math.pi * month / 12),
            "month_cos": math.cos(2 * math.pi * month / 12),
        }
        for value in range(5):
            features[f"dow_{value}"] = float(dow == value)
        return features

    def _make_feature_row(
        self,
        history: pd.Series,
        target_date: Any,
        lags: Sequence[int],
        peak_threshold: float,
    ) -> Dict[str, float]:
        hist = history.dropna().astype(float)
        values = hist.values
        row: Dict[str, float] = {}
        for lag in lags:
            row[f"lag_{lag}"] = float(values[-lag]) if len(values) >= lag else np.nan

        last = float(values[-1]) if len(values) else np.nan
        row["last_value"] = last
        for window in (5, 10, 20, 60):
            recent = values[-window:] if len(values) >= 1 else np.array([])
            row[f"roll_mean_{window}"] = float(np.nanmean(recent)) if len(recent) else np.nan
            row[f"roll_std_{window}"] = float(np.nanstd(recent)) if len(recent) else 0.0
            row[f"roll_max_{window}"] = float(np.nanmax(recent)) if len(recent) else np.nan
            row[f"momentum_{window}"] = float(last - recent[0]) if len(recent) else 0.0

        if np.isfinite(peak_threshold) and len(values):
            row["recent_peak_count_20"] = float(np.sum(values[-20:] >= peak_threshold))
            peak_positions = np.where(values >= peak_threshold)[0]
            row["days_since_peak"] = float(len(values) - 1 - peak_positions[-1]) if len(peak_positions) else 999.0
        else:
            row["recent_peak_count_20"] = 0.0
            row["days_since_peak"] = 999.0

        row.update(self._calendar_features_for_date(target_date))
        return row

    def _create_supervised_frame(
        self,
        series: pd.Series,
        lags: Sequence[int],
        peak_threshold: float,
    ) -> Tuple[pd.DataFrame, pd.Series]:
        clean = series.dropna().astype(float)
        if not lags:
            lags = [1, 2, 3]
        min_start = max(max(lags), 5)
        rows: List[Dict[str, float]] = []
        y_values: List[float] = []
        idx: List[Any] = []

        for i in range(min_start, len(clean)):
            history = clean.iloc[:i]
            rows.append(self._make_feature_row(history, clean.index[i], lags, peak_threshold))
            y_values.append(float(clean.iloc[i]))
            idx.append(clean.index[i])

        X = pd.DataFrame(rows, index=idx)
        y = pd.Series(y_values, index=idx, name="y")
        X = X.replace([np.inf, -np.inf], np.nan)
        X = X.ffill().bfill().fillna(0.0)
        return X, y

    def evaluate_model(
        self,
        actual: Sequence[float],
        predicted: Sequence[float],
        peak_threshold: Optional[float] = None,
    ) -> Dict[str, float]:
        actual_arr = np.asarray(actual, dtype=float)
        pred_arr = np.asarray(predicted, dtype=float)
        mask = np.isfinite(actual_arr) & np.isfinite(pred_arr)
        if mask.sum() < 2:
            return {
                "mae": np.nan,
                "rmse": np.nan,
                "mape": np.nan,
                "smape": np.nan,
                "weighted_mape": np.nan,
                "r2": np.nan,
                "peak_recall": np.nan,
                "peak_mape": np.nan,
            }

        actual_clean = actual_arr[mask]
        pred_clean = pred_arr[mask]
        eps = 1e-8
        denom = np.maximum(np.abs(actual_clean), eps)
        ape = np.abs(actual_clean - pred_clean) / denom
        smape = np.mean(2.0 * np.abs(actual_clean - pred_clean) / np.maximum(np.abs(actual_clean) + np.abs(pred_clean), eps))

        if peak_threshold is None or not np.isfinite(peak_threshold):
            peak_threshold = float(np.nanquantile(actual_clean, self.config.peak_quantile))
        peak_mask = actual_clean >= peak_threshold
        weights = np.where(peak_mask, self.config.peak_weight, 1.0)
        weighted_mape = float(np.average(ape, weights=weights) * 100.0)
        peak_recall = float(np.mean(pred_clean[peak_mask] >= peak_threshold * 0.85)) if peak_mask.any() else np.nan
        peak_mape = float(np.mean(ape[peak_mask]) * 100.0) if peak_mask.any() else np.nan

        try:
            r2 = float(r2_score(actual_clean, pred_clean))
        except Exception:
            r2 = np.nan

        return {
            "mae": float(mean_absolute_error(actual_clean, pred_clean)),
            "rmse": float(np.sqrt(mean_squared_error(actual_clean, pred_clean))),
            "mape": float(np.mean(ape) * 100.0),
            "smape": float(smape * 100.0),
            "weighted_mape": weighted_mape,
            "r2": r2,
            "peak_recall": peak_recall,
            "peak_mape": peak_mape,
        }

    def _fit_arima(self, series: pd.Series, stationary: bool) -> Any:
        order = (1, 0, 1) if stationary else (1, 1, 1)
        return ARIMA(series.dropna().astype(float), order=order).fit()

    def _fit_sarima(self, series: pd.Series, seasonal_period: int, stationary: bool) -> Any:
        d = 0 if stationary else 1
        period = max(2, int(seasonal_period or 5))
        return SARIMAX(
            series.dropna().astype(float),
            order=(1, d, 1),
            seasonal_order=(1, 0, 1, period),
            enforce_stationarity=False,
            enforce_invertibility=False,
        ).fit(disp=False)

    def _fit_ets(self, series: pd.Series, seasonal_period: Optional[int]) -> Any:
        clean = series.dropna().astype(float)
        if seasonal_period and len(clean) >= 2 * seasonal_period:
            model = ExponentialSmoothing(
                clean,
                trend="add",
                damped_trend=True,
                seasonal="add",
                seasonal_periods=int(seasonal_period),
            )
        else:
            model = ExponentialSmoothing(clean, trend="add", damped_trend=True)
        return model.fit(optimized=True)

    def _fit_linear(self, series: pd.Series) -> LinearRegression:
        clean = series.dropna().astype(float)
        x = np.arange(len(clean)).reshape(-1, 1)
        model = LinearRegression()
        model.fit(x, clean.values)
        return model

    def _fit_ml(self, model_name: str, series: pd.Series, lags: Sequence[int], peak_threshold: float) -> Dict[str, Any]:
        X, y = self._create_supervised_frame(series, lags, peak_threshold)
        if len(X) < 40:
            raise ValueError("Not enough supervised rows")

        if model_name == "RANDOM_FOREST":
            model = RandomForestRegressor(
                n_estimators=350,
                min_samples_leaf=3,
                random_state=self.config.random_state,
                n_jobs=-1,
            )
        elif model_name == "GRADIENT_BOOST":
            model = GradientBoostingRegressor(
                n_estimators=250,
                learning_rate=0.04,
                max_depth=3,
                random_state=self.config.random_state,
            )
        elif model_name == "XGBOOST":
            if not self.config.use_xgboost:
                raise ValueError("XGBoost disabled in config")
            try:
                from xgboost import XGBRegressor  # type: ignore
            except Exception as exc:
                raise ImportError("xgboost is not installed") from exc
            model = XGBRegressor(
                n_estimators=350,
                max_depth=3,
                learning_rate=0.035,
                subsample=0.90,
                colsample_bytree=0.90,
                objective="reg:squarederror",
                random_state=self.config.random_state,
                n_jobs=2,
            )
        else:
            raise ValueError(f"Unknown ML model: {model_name}")

        model.fit(X, y)
        return {
            "type": model_name,
            "model": model,
            "lags": list(lags),
            "feature_columns": list(X.columns),
            "peak_threshold": peak_threshold,
        }

    def _apply_peak_lift(
        self,
        forecast: pd.Series,
        reference: pd.Series,
        peak_threshold: float,
    ) -> pd.Series:
        if forecast.empty or not np.isfinite(peak_threshold):
            return forecast
        ref = reference.dropna().astype(float)
        if len(ref) < 60:
            return forecast

        lifted = forecast.copy().astype(float)
        high = ref[ref >= peak_threshold]
        if len(high) < 3:
            return lifted.clip(lower=0) if ref.min() >= 0 else lifted

        for date_value in lifted.index:
            if not isinstance(date_value, pd.Timestamp):
                continue
            same_context = ref.copy()
            if isinstance(ref.index, pd.DatetimeIndex):
                flags = (
                    (ref.index.dayofweek == date_value.dayofweek)
                    | ((ref.index.day >= 25) & (date_value.day >= 25))
                    | ((ref.index.month.isin([3, 6, 9, 12])) & (date_value.month in [3, 6, 9, 12]) & (date_value.day >= 20))
                )
                same_context = ref.loc[flags]
            if len(same_context) < 5:
                continue
            context_q75 = float(same_context.quantile(0.75))
            context_q90 = float(same_context.quantile(0.90))
            high_risk_calendar = date_value.day >= 25 or (date_value.month in [3, 6, 9, 12] and date_value.day >= 20)
            if high_risk_calendar:
                lifted.loc[date_value] = max(float(lifted.loc[date_value]), 0.70 * context_q75 + 0.30 * context_q90)

        return lifted.clip(lower=0) if ref.min() >= 0 else lifted

    def _forecast_model(
        self,
        model_pack: Dict[str, Any],
        history: pd.Series,
        future_index: pd.Index,
        reference_for_peak_lift: Optional[pd.Series] = None,
    ) -> pd.Series:
        model_type = model_pack["type"]
        clean = history.dropna().astype(float)
        horizon = len(future_index)

        if model_type in {"ARIMA", "SARIMA", "EXP_SMOOTHING"}:
            values = np.asarray(model_pack["model"].forecast(steps=horizon), dtype=float)
            forecast = pd.Series(values, index=future_index)
        elif model_type == "LINEAR_REG":
            x_future = np.arange(len(clean), len(clean) + horizon).reshape(-1, 1)
            values = model_pack["model"].predict(x_future)
            forecast = pd.Series(values, index=future_index)
        elif model_type == "NAIVE_LAST":
            forecast = pd.Series(np.repeat(clean.iloc[-1], horizon), index=future_index)
        elif model_type == "SEASONAL_DOW":
            forecast_values = []
            for date_value in future_index:
                if isinstance(date_value, pd.Timestamp) and isinstance(clean.index, pd.DatetimeIndex):
                    same_dow = clean[clean.index.dayofweek == date_value.dayofweek]
                    forecast_values.append(float(same_dow.tail(4).median()) if len(same_dow) else float(clean.iloc[-1]))
                else:
                    forecast_values.append(float(clean.iloc[-1]))
            forecast = pd.Series(forecast_values, index=future_index)
        elif model_type == "KNN_ANALOG_PATH":
            values = model_pack["model"].forecast(clean, horizon=horizon)
            forecast = pd.Series(values, index=future_index)
        elif model_type in {"RANDOM_FOREST", "GRADIENT_BOOST", "XGBOOST"}:
            lags = model_pack["lags"]
            feature_columns = model_pack["feature_columns"]
            peak_threshold = model_pack.get("peak_threshold", np.nan)
            model = model_pack["model"]
            rolling_history = clean.copy()
            preds: List[float] = []
            for date_value in future_index:
                row = self._make_feature_row(rolling_history, date_value, lags, peak_threshold)
                X_row = pd.DataFrame([row]).reindex(columns=feature_columns).replace([np.inf, -np.inf], np.nan)
                X_row = X_row.ffill(axis=1).bfill(axis=1).fillna(0.0)
                pred = float(model.predict(X_row)[0])
                if clean.min() >= 0:
                    pred = max(0.0, pred)
                preds.append(pred)
                rolling_history = pd.concat([rolling_history, pd.Series([pred], index=[date_value])])
            forecast = pd.Series(preds, index=future_index)
        else:
            raise ValueError(f"Unknown model pack type: {model_type}")

        ref = reference_for_peak_lift if reference_for_peak_lift is not None else clean
        return self._apply_peak_lift(forecast, ref, model_pack.get("peak_threshold", self._peak_threshold(ref)))

    def _holdout_length(self, series: pd.Series) -> int:
        n = len(series.dropna())
        holdout = int(round(n * self.config.holdout_ratio))
        holdout = max(self.config.holdout_min, holdout)
        holdout = min(self.config.holdout_max, holdout)
        if n - holdout < self.config.min_train_size:
            holdout = max(10, n - self.config.min_train_size)
        return max(1, min(holdout, n // 3))

    def _baseline_candidates(self, train: pd.Series, test_index: pd.Index, peak_threshold: float) -> List[CandidateResult]:
        candidates: List[CandidateResult] = []
        for name in ("NAIVE_LAST", "SEASONAL_DOW"):
            pack = {"type": name, "model": None, "peak_threshold": peak_threshold}
            pred = self._forecast_model(pack, train, test_index, reference_for_peak_lift=train)
            metrics = self.evaluate_model(self._actual_for_index(test_index), pred.values, peak_threshold)
            candidates.append(CandidateResult(name, None, metrics, pred, pack))
        return candidates

    def _actual_for_index(self, index: pd.Index) -> np.ndarray:
        # This method is patched during candidate evaluation to avoid passing
        # actual arrays through every small helper.
        if not hasattr(self, "_current_test_series"):
            return np.array([])
        return self._current_test_series.reindex(index).values

    def _evaluate_candidate(
        self,
        indicator: str,
        name: str,
        pack: Dict[str, Any],
        train: pd.Series,
        test: pd.Series,
        peak_threshold: float,
    ) -> Optional[CandidateResult]:
        try:
            pred = self._forecast_model(pack, train, test.index, reference_for_peak_lift=train)
            pred = pred.reindex(test.index)
            metrics = self.evaluate_model(test.values, pred.values, peak_threshold)
            if not np.isfinite(metrics.get("weighted_mape", np.nan)):
                return None
            return CandidateResult(name, pack.get("model"), metrics, pred, pack)
        except Exception as exc:
            self._record_failure(indicator, name, str(exc))
            return None

    def fit_best_model(self, series: pd.Series, classification: Dict[str, Any], indicator: str) -> Tuple[Optional[str], Optional[Dict[str, Any]], Dict[str, float]]:
        clean = series.dropna().astype(float)
        if len(clean) < self.config.min_train_size + 10:
            return None, None, {"accuracy_category": "insufficient_data"}

        holdout_len = self._holdout_length(clean)
        train = clean.iloc[:-holdout_len]
        test = clean.iloc[-holdout_len:]
        self._current_test_series = test

        peak_threshold = self._peak_threshold(train)
        selected_lags = self.select_lags(train)
        candidates: List[CandidateResult] = []

        # Baselines first. A complex model must beat them.
        for baseline_name in ("NAIVE_LAST", "SEASONAL_DOW"):
            pack = {"type": baseline_name, "model": None, "peak_threshold": peak_threshold}
            result = self._evaluate_candidate(indicator, baseline_name, pack, train, test, peak_threshold)
            if result:
                candidates.append(result)

        # Parametric models.
        stationary = bool(classification.get("stationarity", {}).get("is_stationary", False))
        seasonal_period = classification.get("seasonality", {}).get("period", 5)
        parametric_builders = [
            ("ARIMA", lambda: {"type": "ARIMA", "model": self._fit_arima(train, stationary), "peak_threshold": peak_threshold}),
            ("SARIMA", lambda: {"type": "SARIMA", "model": self._fit_sarima(train, seasonal_period, stationary), "peak_threshold": peak_threshold}),
            ("EXP_SMOOTHING", lambda: {"type": "EXP_SMOOTHING", "model": self._fit_ets(train, seasonal_period), "peak_threshold": peak_threshold}),
            ("LINEAR_REG", lambda: {"type": "LINEAR_REG", "model": self._fit_linear(train), "peak_threshold": peak_threshold}),
        ]
        for name, builder in parametric_builders:
            try:
                pack = builder()
                result = self._evaluate_candidate(indicator, name, pack, train, test, peak_threshold)
                if result:
                    candidates.append(result)
            except Exception as exc:
                self._record_failure(indicator, name, str(exc))

        # Feature models with leakage-free recursive holdout forecasts.
        for model_name in ("RANDOM_FOREST", "GRADIENT_BOOST", "XGBOOST"):
            try:
                pack = self._fit_ml(model_name, train, selected_lags, peak_threshold)
                result = self._evaluate_candidate(indicator, model_name, pack, train, test, peak_threshold)
                if result:
                    candidates.append(result)
            except Exception as exc:
                self._record_failure(indicator, model_name, str(exc))

        # KNN analog path.
        try:
            analog = KNNAnalogPathModel(
                lookback=min(self.config.analog_lookback, max(10, len(train) // 10)),
                neighbors=self.config.analog_neighbors,
            ).fit(train, horizon=len(test))
            pack = {"type": "KNN_ANALOG_PATH", "model": analog, "peak_threshold": peak_threshold}
            result = self._evaluate_candidate(indicator, "KNN_ANALOG_PATH", pack, train, test, peak_threshold)
            if result:
                candidates.append(result)
        except Exception as exc:
            self._record_failure(indicator, "KNN_ANALOG_PATH", str(exc))

        if not candidates:
            return None, None, {"accuracy_category": "no_model_fitted"}

        comparison_rows = []
        for candidate in candidates:
            row = {"model": candidate.name}
            row.update(candidate.metrics)
            row["selected_lags"] = ",".join(map(str, selected_lags))
            comparison_rows.append(row)
        self.model_comparisons[indicator] = comparison_rows

        baselines = [c for c in candidates if c.name in {"NAIVE_LAST", "SEASONAL_DOW"}]
        best_baseline = min(baselines, key=lambda c: c.metrics["weighted_mape"]) if baselines else None
        best_overall = min(candidates, key=lambda c: c.metrics["weighted_mape"])

        selected = best_overall
        selection_reason = "best_weighted_mape"
        if best_baseline and best_overall.name not in {"NAIVE_LAST", "SEASONAL_DOW"}:
            baseline_score = best_baseline.metrics["weighted_mape"]
            model_score = best_overall.metrics["weighted_mape"]
            gain = (baseline_score - model_score) / baseline_score if baseline_score > 0 else 0.0
            if gain < self.config.complex_model_min_gain:
                selected = best_baseline
                selection_reason = "baseline_kept_complex_gain_below_5pct"
        elif best_baseline and best_overall.name in {"NAIVE_LAST", "SEASONAL_DOW"}:
            selection_reason = "baseline_won_holdout"

        # Refit selected model on all data for future projection. If the final
        # refit fails, keep the run alive and fall back to the strongest
        # baseline instead of exporting a broken forecast.
        try:
            final_pack = self._refit_selected_model(selected.name, clean, selected_lags, classification)
        except Exception as exc:
            self._record_failure(indicator, f"{selected.name}_FINAL_REFIT", str(exc))
            if best_baseline is not None:
                selected = best_baseline
                final_pack = {
                    "type": selected.name,
                    "model": None,
                    "peak_threshold": self._peak_threshold(clean),
                }
                selection_reason = "fallback_baseline_after_final_refit_failure"
            else:
                return None, None, {"accuracy_category": "final_refit_failed"}
        final_pack["holdout_metrics"] = selected.metrics
        final_pack["holdout_predictions"] = selected.predictions
        final_pack["selected_lags"] = selected_lags
        final_pack["selection_reason"] = selection_reason
        final_pack["baseline_model"] = best_baseline.name if best_baseline else None
        final_pack["baseline_weighted_mape"] = best_baseline.metrics["weighted_mape"] if best_baseline else np.nan
        final_pack["peak_threshold"] = final_pack.get("peak_threshold", self._peak_threshold(clean))

        metrics = selected.metrics.copy()
        metrics["accuracy_category"] = self._accuracy_category(metrics)
        metrics["selection_reason"] = selection_reason
        metrics["baseline_weighted_mape"] = final_pack["baseline_weighted_mape"]

        self.holdout_predictions[indicator] = selected.predictions
        return selected.name, final_pack, metrics

    def _refit_selected_model(
        self,
        model_name: str,
        series: pd.Series,
        selected_lags: Sequence[int],
        classification: Dict[str, Any],
    ) -> Dict[str, Any]:
        peak_threshold = self._peak_threshold(series)
        stationary = bool(classification.get("stationarity", {}).get("is_stationary", False))
        seasonal_period = classification.get("seasonality", {}).get("period", 5)

        if model_name == "ARIMA":
            return {"type": "ARIMA", "model": self._fit_arima(series, stationary), "peak_threshold": peak_threshold}
        if model_name == "SARIMA":
            return {"type": "SARIMA", "model": self._fit_sarima(series, seasonal_period, stationary), "peak_threshold": peak_threshold}
        if model_name == "EXP_SMOOTHING":
            return {"type": "EXP_SMOOTHING", "model": self._fit_ets(series, seasonal_period), "peak_threshold": peak_threshold}
        if model_name == "LINEAR_REG":
            return {"type": "LINEAR_REG", "model": self._fit_linear(series), "peak_threshold": peak_threshold}
        if model_name in {"RANDOM_FOREST", "GRADIENT_BOOST", "XGBOOST"}:
            return self._fit_ml(model_name, series, selected_lags, peak_threshold)
        if model_name == "KNN_ANALOG_PATH":
            analog = KNNAnalogPathModel(
                lookback=min(self.config.analog_lookback, max(10, len(series) // 10)),
                neighbors=self.config.analog_neighbors,
            ).fit(series, horizon=self.horizon)
            return {"type": "KNN_ANALOG_PATH", "model": analog, "peak_threshold": peak_threshold}
        if model_name in {"NAIVE_LAST", "SEASONAL_DOW"}:
            return {"type": model_name, "model": None, "peak_threshold": peak_threshold}
        raise ValueError(f"Unsupported selected model: {model_name}")

    @staticmethod
    def _accuracy_category(metrics: Dict[str, float]) -> str:
        wmape = metrics.get("weighted_mape", np.nan)
        if not np.isfinite(wmape):
            return "unknown"
        if wmape <= 5:
            return "excellent"
        if wmape <= 10:
            return "good"
        if wmape <= 20:
            return "moderate"
        return "weak"

    def forecast_series(self, series: pd.Series, model_pack: Dict[str, Any]) -> Optional[pd.Series]:
        try:
            clean = series.dropna().astype(float)
            if len(clean) == 0:
                return None
            if isinstance(clean.index, pd.DatetimeIndex):
                future_index = make_future_business_dates(clean.index[-1], self.horizon, BASE_DIR)
            else:
                future_index = pd.RangeIndex(len(clean), len(clean) + self.horizon)
            return self._forecast_model(model_pack, clean, future_index, reference_for_peak_lift=clean)
        except Exception as exc:
            print("FORECAST FAILED:", type(exc).__name__, exc)
            return None

    def analyze_correlations(self) -> Optional[Dict[str, Any]]:
        df_numeric = self.df[self.indicator_cols].select_dtypes(include=[np.number])
        if df_numeric.empty:
            return None
        corr_matrix = df_numeric.corr()
        strong_corr = []
        for i in range(len(corr_matrix.columns)):
            for j in range(i + 1, len(corr_matrix.columns)):
                corr_val = corr_matrix.iloc[i, j]
                if np.isfinite(corr_val) and abs(corr_val) > 0.7:
                    strong_corr.append(
                        {
                            "indicator_1": corr_matrix.columns[i],
                            "indicator_2": corr_matrix.columns[j],
                            "correlation": corr_val,
                            "relationship": "positive" if corr_val > 0 else "negative",
                        }
                    )
        return {"correlation_matrix": corr_matrix, "strong_correlations": strong_corr}

    def analyze_all(self) -> None:
        print("=" * 80)
        print("ROBUST TIME-SERIES FORECASTING PIPELINE")
        print("=" * 80)

        self.df = self.df.sort_index()
        print("No calendar interpolation is applied. Closed days remain absent.")

        for indicator in self.indicator_cols:
            if indicator not in self.df.columns:
                continue
            print(f"\nAnalyse de: {indicator}")
            series = pd.to_numeric(self.df[indicator], errors="coerce").dropna()
            if series.empty:
                print("  skipped: empty numeric series")
                continue

            classification = self.classify_series(series, indicator)
            self.classifications[indicator] = classification
            print(f"  Classification: {classification['classification']}")

            try:
                model_name, model_pack, metrics = self.fit_best_model(series, classification, indicator)
            except Exception as exc:
                self._record_failure(indicator, "SERIES_FAILED", str(exc))
                print(f"  Series failed and was skipped: {type(exc).__name__}: {exc}")
                continue
            if model_pack:
                self.models[indicator] = {"name": model_name, "model": model_pack, "metrics": metrics}
                print(
                    "  Best model: "
                    f"{model_name} | WMAPE={metrics.get('weighted_mape', np.nan):.3f}% "
                    f"| MAPE={metrics.get('mape', np.nan):.3f}% "
                    f"| reason={metrics.get('selection_reason')}"
                )
                forecast = self.forecast_series(series, model_pack)
                if forecast is not None:
                    self.forecasts[indicator] = forecast
                    print(f"  Projection generated on {len(forecast)} Moroccan business days")
            else:
                print(f"  No model fitted: {metrics.get('accuracy_category')}")

        print("\nAnalyse des dependances...")
        self.results["correlations"] = self.analyze_correlations()
        self.results["failures"] = self.failures
        print("=" * 80)
        print("ANALYSIS COMPLETE")


class ResultsExporter:
    """Excel and HTML exports, keeping the original visual style."""

    def __init__(self, analyzer: TimeSeriesAnalyzer, output_dir: Path | str) -> None:
        self.analyzer = analyzer
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def create_summary_report(self) -> Path:
        wb = Workbook()
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws_summary = wb.active
        ws_summary.title = "Resume Classifications"
        headers = [
            "Indicateur",
            "Classification",
            "Tendance",
            "Saisonnalite",
            "Volatilite",
            "Modele",
            "MAPE",
            "WMAPE peak-aware",
            "Peak recall",
            "Baseline WMAPE",
            "Raison selection",
            "Lags selectionnes",
        ]
        ws_summary.append(headers)
        self._style_header(ws_summary, header_fill, header_font)

        for indicator in self.analyzer.indicator_cols:
            classif = self.analyzer.classifications.get(indicator)
            model_info = self.analyzer.models.get(indicator, {})
            if not classif:
                continue
            metrics = model_info.get("metrics", {})
            model_pack = model_info.get("model", {})
            ws_summary.append(
                [
                    indicator,
                    classif["classification"],
                    classif["trend"]["trend_type"],
                    "Oui" if classif["seasonality"]["has_seasonality"] else "Non",
                    classif["volatility"]["volatility_type"],
                    model_info.get("name", "N/A"),
                    metrics.get("mape", "N/A"),
                    metrics.get("weighted_mape", "N/A"),
                    metrics.get("peak_recall", "N/A"),
                    metrics.get("baseline_weighted_mape", "N/A"),
                    metrics.get("selection_reason", "N/A"),
                    ",".join(map(str, model_pack.get("selected_lags", []))),
                ]
            )

        ws_metrics = wb.create_sheet("Comparaison Modeles")
        ws_metrics.append(
            [
                "Indicateur",
                "Modele",
                "MAE",
                "RMSE",
                "MAPE",
                "sMAPE",
                "WMAPE peak-aware",
                "R2 indicatif",
                "Peak recall",
                "Peak MAPE",
                "Lags",
            ]
        )
        self._style_header(ws_metrics, header_fill, header_font)
        for indicator, rows in self.analyzer.model_comparisons.items():
            for row in rows:
                ws_metrics.append(
                    [
                        indicator,
                        row.get("model"),
                        row.get("mae"),
                        row.get("rmse"),
                        row.get("mape"),
                        row.get("smape"),
                        row.get("weighted_mape"),
                        row.get("r2"),
                        row.get("peak_recall"),
                        row.get("peak_mape"),
                        row.get("selected_lags"),
                    ]
                )

        ws_forecast = wb.create_sheet("Projections")
        if self.analyzer.forecasts:
            all_dates = sorted({d for forecast in self.analyzer.forecasts.values() for d in forecast.index})
            ws_forecast.append(["Indicateur"] + all_dates)
            self._style_header(ws_forecast, header_fill, header_font)
            for indicator, forecast in self.analyzer.forecasts.items():
                ws_forecast.append([indicator] + [forecast.get(d, None) for d in all_dates])

        ws_diag = wb.create_sheet("Diagnostics")
        ws_diag.append(["cle", "valeur"])
        self._style_header(ws_diag, header_fill, header_font)
        ws_diag.append(["n_series", len(self.analyzer.indicator_cols)])
        ws_diag.append(["n_models", len(self.analyzer.models)])
        ws_diag.append(["n_failures", len(self.analyzer.failures)])
        ws_diag.append(["config", json.dumps(self.analyzer.config.__dict__, ensure_ascii=False)])
        for failure in self.analyzer.failures[:200]:
            ws_diag.append([f"failure::{failure['indicator']}::{failure['model']}", failure["reason"]])

        if self.analyzer.results.get("correlations"):
            ws_corr = wb.create_sheet("Correlations")
            ws_corr.append(["Indicateur 1", "Indicateur 2", "Correlation", "Type"])
            self._style_header(ws_corr, header_fill, header_font)
            for corr in self.analyzer.results["correlations"]["strong_correlations"]:
                ws_corr.append([corr["indicator_1"], corr["indicator_2"], corr["correlation"], corr["relationship"]])

        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 70)
                for cell in column:
                    cell.alignment = Alignment(vertical="top")

        output_path = self.output_dir / "analyse_series_chronologiques_robuste.xlsx"
        wb.save(output_path)
        print(f"Excel report saved: {output_path}")
        return output_path

    @staticmethod
    def _style_header(ws: Any, fill: PatternFill, font: Font) -> None:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    def create_visualization_dashboard(self) -> None:
        if not self.analyzer.classifications:
            return

        class_counts: Dict[str, int] = {}
        for classif in self.analyzer.classifications.values():
            cls = classif["classification"]
            class_counts[cls] = class_counts.get(cls, 0) + 1

        fig_class = go.Figure(
            data=[go.Pie(labels=list(class_counts.keys()), values=list(class_counts.values()), hole=0.35)]
        )
        fig_class.update_layout(title="Distribution des classifications", template="plotly_white")
        fig_class.write_html(self.output_dir / "classification_distribution.html")

        model_names = []
        weighted_mapes = []
        for indicator, model_info in self.analyzer.models.items():
            wmape = model_info.get("metrics", {}).get("weighted_mape", np.nan)
            if np.isfinite(wmape):
                model_names.append(f"{indicator[:35]}...")
                weighted_mapes.append(wmape)

        if weighted_mapes:
            fig_acc = go.Figure(data=[go.Bar(x=model_names, y=weighted_mapes, marker_color="#1F4E78")])
            fig_acc.update_layout(
                title="Qualite holdout des modeles (WMAPE peak-aware, plus bas = mieux)",
                xaxis_tickangle=-45,
                yaxis_title="WMAPE (%)",
                template="plotly_white",
                height=650,
            )
            fig_acc.write_html(self.output_dir / "model_accuracy.html")

        corr = self.analyzer.results.get("correlations") if self.analyzer.results else None
        if corr and corr.get("correlation_matrix") is not None:
            corr_matrix = corr["correlation_matrix"]
            fig_corr = go.Figure(
                data=go.Heatmap(
                    z=corr_matrix.values,
                    x=corr_matrix.columns,
                    y=corr_matrix.columns,
                    colorscale="RdBu",
                    zmid=0,
                )
            )
            fig_corr.update_layout(title="Matrice de correlation", template="plotly_white", height=800)
            fig_corr.write_html(self.output_dir / "correlation_matrix.html")

        print("Dashboard charts created.")

    def create_individual_charts(self) -> None:
        for indicator in self.analyzer.indicator_cols:
            if indicator not in self.analyzer.classifications:
                continue
            series = pd.to_numeric(self.analyzer.df[indicator], errors="coerce").dropna()
            classif = self.analyzer.classifications[indicator]
            model_info = self.analyzer.models.get(indicator, {})
            model_pack = model_info.get("model", {})
            holdout_pred = self.analyzer.holdout_predictions.get(indicator)
            forecast = self.analyzer.forecasts.get(indicator)
            peak_threshold = classif.get("peaks", {}).get("threshold", np.nan)

            fig = make_subplots(
                rows=2,
                cols=1,
                subplot_titles=(f"{indicator} - historique, holdout et projection", "Residus holdout"),
                vertical_spacing=0.15,
                row_heights=[0.72, 0.28],
            )
            fig.add_trace(
                go.Scatter(
                    x=series.index,
                    y=series.values,
                    mode="lines+markers",
                    name="Historique",
                    line=dict(color="#1F4E78", width=2),
                ),
                row=1,
                col=1,
            )

            if np.isfinite(peak_threshold):
                peaks = series[series >= peak_threshold]
                if not peaks.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=peaks.index,
                            y=peaks.values,
                            mode="markers",
                            name="Pics historiques",
                            marker=dict(color="#E67E22", size=8, symbol="diamond"),
                        ),
                        row=1,
                        col=1,
                    )

            if holdout_pred is not None:
                fig.add_trace(
                    go.Scatter(
                        x=holdout_pred.index,
                        y=holdout_pred.values,
                        mode="lines+markers",
                        name="Prediction holdout",
                        line=dict(color="#27AE60", width=2, dash="dot"),
                    ),
                    row=1,
                    col=1,
                )
                actual_holdout = series.reindex(holdout_pred.index)
                residuals = actual_holdout - holdout_pred
                fig.add_trace(
                    go.Scatter(
                        x=residuals.index,
                        y=residuals.values,
                        mode="markers",
                        name="Residus vrais",
                        marker=dict(color="#95A5A6", size=5),
                    ),
                    row=2,
                    col=1,
                )
                fig.add_hline(y=0, line_dash="dash", line_color="gray", row=2, col=1)

            if forecast is not None:
                fig.add_trace(
                    go.Scatter(
                        x=forecast.index,
                        y=forecast.values,
                        mode="lines+markers",
                        name="Projection",
                        line=dict(color="#FF6B6B", width=2, dash="dash"),
                    ),
                    row=1,
                    col=1,
                )

            selected_lags = ",".join(map(str, model_pack.get("selected_lags", [])))
            wmape = model_info.get("metrics", {}).get("weighted_mape", np.nan)
            title = (
                f"{indicator}<br><sub>"
                f"Classification: {classif['classification']} | "
                f"Modele: {model_info.get('name', 'N/A')} | "
                f"WMAPE: {wmape:.3f}% | Lags: {selected_lags[:80]}"
                "</sub>"
            )
            fig.update_layout(title=title, template="plotly_white", height=850, showlegend=True, hovermode="x unified")
            fig.update_xaxes(title_text="Date", row=2, col=1)
            fig.update_yaxes(title_text="Valeur", row=1, col=1)
            fig.update_yaxes(title_text="Erreur", row=2, col=1)

            safe_name = re.sub(r"[^\w\-_\. ]", "_", indicator)[:90]
            fig.write_html(self.output_dir / f"serie_{safe_name}.html")

        print(f"Individual charts created for {len(self.analyzer.indicator_cols)} indicators.")


def load_excel_data(input_dir: Path | str) -> pd.DataFrame:
    input_path = Path(input_dir)
    excel_files = sorted([p for p in input_path.glob("*.xlsx") if not p.name.startswith("~$")])
    excel_files.extend(sorted([p for p in input_path.glob("*.xls") if not p.name.startswith("~$")]))
    if not excel_files:
        raise FileNotFoundError(f"No Excel file found in {input_path}")
    selected = excel_files[0]
    df = pd.read_excel(selected)
    print(f"Loaded {selected.name}: {df.shape[0]} rows, {df.shape[1]} columns")
    return df


def detect_date_column(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        normalized = normalize_column_name(col)
        if "date" in normalized or "jour" in normalized or "day" in normalized:
            return col
    for col in df.columns:
        parsed = pd.to_datetime(df[col], errors="coerce")
        if parsed.notna().mean() > 0.80:
            return col
    if len(df.columns) and not pd.api.types.is_numeric_dtype(df.iloc[:, 0]):
        return df.columns[0]
    return None


def detect_indicator_columns(df: pd.DataFrame, date_col: Optional[str]) -> List[str]:
    cols = []
    for col in df.columns:
        if col == date_col:
            continue
        numeric = pd.to_numeric(df[col], errors="coerce")
        if numeric.notna().mean() > 0.80:
            df[col] = numeric
            cols.append(col)
    return cols


def main() -> None:
    print("=" * 80)
    print("ROBUST FORECASTING PIPELINE")
    print("=" * 80)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    df = load_excel_data(INPUT_DIR)
    date_col = detect_date_column(df)
    print(f"Date column: {date_col}")
    indicator_cols = detect_indicator_columns(df, date_col)
    print(f"{len(indicator_cols)} numeric indicators detected")

    analyzer = TimeSeriesAnalyzer(df, date_col=date_col, indicator_cols=indicator_cols)
    analyzer.analyze_all()

    exporter = ResultsExporter(analyzer, OUTPUT_DIR)
    exporter.create_summary_report()
    exporter.create_visualization_dashboard()
    exporter.create_individual_charts()

    print("=" * 80)
    print(f"All outputs are available in: {OUTPUT_DIR}")
    print("=" * 80)


if __name__ == "__main__":
    main()
