"""
Pipeline pour analyse des series chrono:
(Exemple: Les composants du pnb pour les dépots et les crédits des TOP 100 clients sur base de l encours moyen)
Analyse, classification, modélisation et projection avec évaluation de précision
"""

import pandas as pd
import numpy as np
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Statistical and ML libraries
from scipy import stats
from scipy.signal import find_peaks
from statsmodels.tsa.seasonal import seasonal_decompose
from statsmodels.tsa.stattools import adfuller, acf, pacf
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from statsmodels.tsa.vector_ar.var_model import VAR
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score



# Visualization
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

# Paths
INPUT_DIR = Path("C:\work\PROJETS_MODELISATION\Analyses Mr Hamlili\Historique PNB dépots et crédits\Séries_chrono\pl_timeseries_pipeline\in")
OUTPUT_DIR = Path("C:\work\PROJETS_MODELISATION\Analyses Mr Hamlili\Historique PNB dépots et crédits\Séries_chrono\pl_timeseries_pipeline\out")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


class TimeSeriesAnalyzer:
    """Analyseur de séries chronologiques avec classification et modélisation"""
    
    def __init__(self, df, date_col=None, indicator_cols=None):
        self.df = df.copy()
        self.date_col = date_col
        self.indicator_cols = indicator_cols or [col for col in df.columns if col != date_col]
        self.horizon = 30
        
        # Préparer les données
        if date_col:
            self.df[date_col] = pd.to_datetime(self.df[date_col])
            self.df = self.df.sort_values(date_col)
            self.df.set_index(date_col, inplace=True)
        
        self.results = {}
        self.models = {}
        self.forecasts = {}
        self.classifications = {}
        
    def analyze_stationarity(self, series):
        """Test de stationnarité (ADF test)"""
        result = adfuller(series.dropna())
        return {
            'adf_statistic': result[0],
            'p_value': result[1],
            'is_stationary': result[1] < 0.05,
            'critical_values': result[4]
        }
    
    def detect_trend(self, series):
        """Détection de tendance"""
        x = np.arange(len(series))
        y = series.values
        
        # Régression linéaire
        mask = ~np.isnan(y)
        if mask.sum() < 2:
            return {'trend_type': 'insufficient_data', 'slope': 0, 'r2': 0}
        
        slope, intercept, r_value, p_value, std_err = stats.linregress(x[mask], y[mask])
        
        # Classification de la tendance
        if p_value > 0.05:
            trend_type = 'no_trend'
        elif slope > 0:
            trend_type = 'increasing'
        else:
            trend_type = 'decreasing'
        
        return {
            'trend_type': trend_type,
            'slope': slope,
            'intercept': intercept,
            'r2': r_value**2,
            'p_value': p_value
        }
    
    def detect_seasonality(self, series, period=None):
        """Détection de saisonnalité"""
        if len(series.dropna()) < 2 * (period or 30):
            return {'has_seasonality': False, 'strength': 0}
        
        try:
            # Décomposition
            if period is None:
                period = min(30, len(series) // 2)
            
            decomposition = seasonal_decompose(series.dropna(), period=period, extrapolate_trend='freq')
            
            # Force de la saisonnalité
            seasonal_strength = np.var(decomposition.seasonal) / np.var(series.dropna())
            
            return {
                'has_seasonality': seasonal_strength > 0.1,
                'strength': seasonal_strength,
                'period': period
            }
        except:
            return {'has_seasonality': False, 'strength': 0}
    
    def detect_volatility(self, series):
        """Analyse de volatilité"""
        returns = series.pct_change().dropna()
        
        if len(returns) == 0:
            return {'volatility': 0, 'volatility_type': 'stable'}
        
        volatility = returns.std()
        cv = abs(returns.std() / returns.mean()) if returns.mean() != 0 else np.inf
        
        # Classification
        if cv < 0.5:
            vol_type = 'low'
        elif cv < 1.5:
            vol_type = 'medium'
        else:
            vol_type = 'high'
        
        return {
            'volatility': volatility,
            'coefficient_variation': cv,
            'volatility_type': vol_type,
            'max_drawdown': (series / series.cummax() - 1).min() if len(series) > 0 else 0
        }
    
    def classify_series(self, series, indicator_name):
        """Classification du comportement de la série"""
        stationarity = self.analyze_stationarity(series)
        trend = self.detect_trend(series)
        seasonality = self.detect_seasonality(series)
        volatility = self.detect_volatility(series)
        
        # Classification principale
        if trend['trend_type'] == 'no_trend' and not seasonality['has_seasonality']:
            if volatility['volatility_type'] == 'low':
                classification = 'STABLE'
            else:
                classification = 'RANDOM_WALK'
        elif trend['trend_type'] != 'no_trend' and seasonality['has_seasonality']:
            classification = 'TREND_SEASONAL'
        elif trend['trend_type'] != 'no_trend':
            classification = 'TRENDING'
        elif seasonality['has_seasonality']:
            classification = 'SEASONAL'
        else:
            classification = 'COMPLEX'
        
        return {
            'indicator': indicator_name,
            'classification': classification,
            'stationarity': stationarity,
            'trend': trend,
            'seasonality': seasonality,
            'volatility': volatility
        }
    
    def fit_arima_model(self, series, order=(1,1,1), auto=False):
         """Ajustement modèle ARIMA"""
         try:
             if auto:
                 from pmdarima import auto_arima
                 fitted_model = auto_arima(series.dropna(), seasonal=False, stepwise=True, suppress_warnings=True)
                 return fitted_model
             else:
                 model = ARIMA(series.dropna(), order=order)
                 fitted_model = model.fit()
                 return fitted_model
         except:
             return None
         
    def fit_sarima_model(self,series,order=(1,1,1),seasonal_order=(1,1,1,30)):
       
       """Ajustement modèle SARIMA"""
       try:
           model = SARIMAX(
               series.dropna(),
               order=order,
               seasonal_order=seasonal_order,
               enforce_stationarity=False,
               enforce_invertibility=False
           )
           fitted_model = model.fit(disp=False)
           return fitted_model
       except:
           return None

    
    def fit_exponential_smoothing(self, series, seasonal_periods=None):
        """Ajustement modèle Exponential Smoothing"""
        try:
            if seasonal_periods and len(series.dropna()) >= 2 * seasonal_periods:
                model = ExponentialSmoothing(
                    series.dropna(), 
                    seasonal_periods=seasonal_periods,
                    trend='add',
                    seasonal='add',
                    damped_trend=True,
                )
            else:
                model = ExponentialSmoothing(series.dropna(), trend='add',damped_trend=True)
            
            fitted_model = model.fit()
            return fitted_model
        except:
            return None
        
    
    
    def fit_linear_regression(self, series):
        """Ajustement régression linéaire"""
        x = np.arange(len(series)).reshape(-1, 1)
        y = series.values
        
        mask = ~np.isnan(y)
        if mask.sum() < 2:
            return None
        
        model = LinearRegression()
        model.fit(x[mask], y[mask])
        return model
    
    def fit_random_forest(self, series, n_lags=90):
       """Ajustement Random Forest avec retards"""
       try:
           df = self._create_lag_features(series, n_lags)
           if len(df) < 10:
               return None
           X = df.drop("y", axis=1)
           y = df["y"]
           model = RandomForestRegressor(
               n_estimators=300,
               random_state=42,
               n_jobs=-1
           )
           model.fit(X, y)
           return model
       except:
           return None
    
    def fit_gradient_boosting(self, series, n_lags=90):
       """Ajustement Gradient Boosting avec retards"""
       try:
           df = self._create_lag_features(series, n_lags)
           if len(df) < 10:
               return None
           X = df.drop("y", axis=1)
           y = df["y"]
           model = GradientBoostingRegressor(random_state=42)
           model.fit(X, y)
           return model
       except:
           return None
       
     # Features Lags#
     #   
    def _create_lag_features(self, series, n_lags):
       df = series.to_frame("y").copy()
       for lag in range(1, n_lags + 1):
           df[f"lag_{lag}"] = df["y"].shift(lag)
       return df.dropna()
    
 
    
    def evaluate_model(self, actual, predicted):
        """Évaluation de la précision du modèle"""
        mask = ~(np.isnan(actual) | np.isnan(predicted))
        
        if mask.sum() < 2:
            return {
                'mae': np.nan,
                'rmse': np.nan,
                'mape': np.nan,
                'r2': np.nan,
                'accuracy_category': 'insufficient_data'
            }
        
        actual_clean = actual[mask]
        predicted_clean = predicted[mask]
        
        mae = mean_absolute_error(actual_clean, predicted_clean)
        rmse = np.sqrt(mean_squared_error(actual_clean, predicted_clean))
        
        # MAPE (éviter division par zéro)
        mape = np.mean(np.abs((actual_clean - predicted_clean) / actual_clean)) * 100
        mape = mape if not np.isinf(mape) else np.nan
        
        r2 = r2_score(actual_clean, predicted_clean)
        
        # Catégorisation de la précision
        if r2 > 0.9:
            accuracy_cat = 'excellent'
        elif r2 > 0.7:
            accuracy_cat = 'good'
        elif r2 > 0.5:
            accuracy_cat = 'moderate'
        else:
            accuracy_cat = 'poor'
        
        return {
            'mae': mae,
            'rmse': rmse,
            'mape': mape,
            'r2': r2,
            'accuracy_category': accuracy_cat
        }
    
    def fit_best_model(self, series, classification):
        """Sélection et ajustement du meilleur modèle selon la classification"""
        models_tested = []
        
        # Split train/test (80/20)
        split_idx = int(len(series) * 0.8)
        train = series[:split_idx]
        test = series[split_idx:]
        
        if len(train) < 3 or len(test) < 1:
            return None, None, {'accuracy_category': 'insufficient_data'}
        
        # Modèle 1: ARIMA
        try:
            arima_model = self.fit_arima_model(train)
            if arima_model:
                arima_pred = arima_model.forecast(steps=len(test))
                arima_metrics = self.evaluate_model(test.values, arima_pred)
                models_tested.append(('ARIMA', arima_model, arima_metrics))
        except:
            pass

        # Modèle 2: SARIMA
        try:
             seasonal_period = classification.get('seasonality', {}).get('period', 30)
             sarima_model = self.fit_sarima_model(train,order=(1,1,1),seasonal_order=(1,1,1,seasonal_period))
             if sarima_model:
                 pred = sarima_model.forecast(steps=len(test))
                 metrics = self.evaluate_model(test.values, pred)
                 models_tested.append(('SARIMA', sarima_model, metrics))
        except:
            pass

        # Modèle 3: Exponential Smoothing
        try:
            seasonal_period = classification.get('seasonality', {}).get('period', 30)
            es_model = self.fit_exponential_smoothing(train, seasonal_period)
            if es_model:
                es_pred = es_model.forecast(steps=len(test))
                es_metrics = self.evaluate_model(test.values, es_pred)
                models_tested.append(('EXP_SMOOTHING', es_model, es_metrics))
        except:
            pass
        
        # Modèle 4: Régression linéaire
        try:
            lr_model = self.fit_linear_regression(train)
            if lr_model:
                x_test = np.arange(len(train), len(series)).reshape(-1, 1)
                lr_pred = lr_model.predict(x_test)
                lr_metrics = self.evaluate_model(test.values, lr_pred)
                models_tested.append(('LINEAR_REG', lr_model, lr_metrics))
        except:
            pass

        # Modèle 5 Random Forest
        try:
            rf_model = self.fit_random_forest(series)
            if rf_model:
               df_feat = self._create_lag_features(series, n_lags=90)
               split_ml = int(len(df_feat) * 0.8)
               test_ml = df_feat[split_ml:]
               pred = rf_model.predict(test_ml.drop("y", axis=1))
               metrics = self.evaluate_model(test_ml["y"].values, pred)
               models_tested.append(('RANDOM_FOREST', rf_model, metrics))
        except:
            pass

        # Modèle 6 Gradient Boosting
        try:
             gb_model = self.fit_gradient_boosting(series)
             if gb_model:
                df_feat = self._create_lag_features(series, n_lags=90)
                split_ml = int(len(df_feat) * 0.8)
                test_ml = df_feat[split_ml:]
                pred = gb_model.predict(test_ml.drop("y", axis=1))
                metrics = self.evaluate_model(test_ml["y"].values, pred)
                models_tested.append(('GRADIENT_BOOST', gb_model, metrics))
        except:
             pass
    
        # Sélectionner le meilleur modèle
        if not models_tested:
            return None, None, {'accuracy_category': 'no_model_fitted'}
        
        best_model = max(models_tested, key=lambda x: x[2].get('r2', -np.inf) if not np.isnan(x[2].get('r2', -np.inf)) else -np.inf)
        
        return best_model[0], best_model[1], best_model[2]
    
    def forecast_series(self, series, model_name, model):
        """Projection future de la série"""
        try:
            series = series.dropna()
            forecast = None
            if model_name == 'ARIMA':
                forecast = model.forecast(steps=self.horizon)
            elif model_name == 'SARIMA':
                forecast = model.forecast(steps=self.horizon)
            elif model_name == 'EXP_SMOOTHING':
                forecast = model.forecast(steps=self.horizon)
            elif model_name == 'LINEAR_REG':
                x_future = np.arange(len(series), len(series) + self.horizon).reshape(-1, 1)
                forecast = model.predict(x_future)
            elif model_name in ['RANDOM_FOREST', 'GRADIENT_BOOST']:
                # Forecast récursif basé sur les lags
                n_lags = model.n_features_in_
                # Nettoyage serie
                series = series.dropna()
                #test len series
                if len(series)<n_lags:
                    print(f"Series trop courte pour {model_name}: {len(series)}<{n_lags}")
                    return None
                last_values = series.values[-n_lags:].tolist()
                forecast = []
                for _ in range(self.horizon):
                    X = np.array(last_values[-n_lags:]).reshape(1, -1)
                    y_pred = model.predict(X)[0]
                    forecast.append(y_pred)
                    last_values.append(y_pred)
                forecast = np.array(forecast)
            else:
                print("UNKNOWN MODEL TYPE")
                return None
            
            # Créer l'index de dates futures
            future_dates = None
            if len(series) > 0:
                last_date = series.index[-1]
                if isinstance(last_date, pd.Timestamp):
                    freq = pd.infer_freq(series.index)
                    if freq is None:
                        freq = 'D'
                    future_dates = pd.date_range(start=last_date,periods=self.horizon + 1,freq=freq)[1:]
                else:
                    future_dates = range(len(series), len(series) + self.horizon)
            if future_dates is None:
                future_dates = range(self.horizon)
            return pd.Series(forecast, index=future_dates)
            
        except Exception as e:
            print("FORECAST FAILED WITH ERROR:", type(e), e)
            import traceback
            traceback.print_exc()
            return None
    
    def analyze_correlations(self):
        """Analyse des corrélations et dépendances entre indicateurs"""
        df_numeric = self.df[self.indicator_cols].select_dtypes(include=[np.number])
        
        if df_numeric.empty:
            return None
        
        # Matrice de corrélation
        corr_matrix = df_numeric.corr()
        
        # Identifier les fortes corrélations (> 0.7 ou < -0.7)
        strong_corr = []
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                corr_val = corr_matrix.iloc[i, j]
                if abs(corr_val) > 0.7:
                    strong_corr.append({
                        'indicator_1': corr_matrix.columns[i],
                        'indicator_2': corr_matrix.columns[j],
                        'correlation': corr_val,
                        'relationship': 'positive' if corr_val > 0 else 'negative'
                    })
        
        return {
            'correlation_matrix': corr_matrix,
            'strong_correlations': strong_corr
        }
    
    def analyze_all(self):
        """Analyse complète de toutes les séries"""
        print("=" * 80)
        print("ANALYSE DES SÉRIES CHRONOLOGIQUES")
        print("=" * 80)
        
        # S'assurer que l'index est datetime avec fréquence
        if not pd.api.types.is_datetime64_any_dtype(self.df.index):
            self.df.index = pd.to_datetime(self.df.index)
        self.df = self.df.sort_index()
        if self.df.index.freq is None:
            self.df = self.df.asfreq("D")  # ou "M" si données mensuelles
        self.df = self.df.interpolate()  # combler éventuels trous

        for indicator in self.indicator_cols:
            print(f"\n Analyse de: {indicator}")
            
            series = self.df[indicator].dropna()
            
            # Classification
            classification = self.classify_series(series, indicator)
            self.classifications[indicator] = classification
            print(f" Classification: {classification['classification']}")
            
            # Modélisation
            print(f" Ajustement des modèles...")
            model_name, model, metrics = self.fit_best_model(series, classification)
            
            if model:
                self.models[indicator] = {
                    'name': model_name,
                    'model': model,
                    'metrics': metrics
                }
                print(f" Meilleur modèle: {model_name}")
                print(f" Précision (R²): {metrics.get('r2', 0):.3f} - {metrics['accuracy_category']}")
                
                # Projection
                forecast = self.forecast_series(series, model_name, model)
                if forecast is not None:
                    self.forecasts[indicator] = forecast
                    print(f"  Projection réalisée sur {self.horizon} périodes")
                else:
                    print(f"Forecast impossible pour {indicator}")
            else:
                print(f" Aucun modèle n'a pu être ajusté")
        
        # Analyse des corrélations
        print(f"\n Analyse des dépendances...")
        corr_analysis = self.analyze_correlations()
        if corr_analysis and corr_analysis['strong_correlations']:
            print(f" {len(corr_analysis['strong_correlations'])} corrélations fortes identifiées")
        
        self.results['correlations'] = corr_analysis
        
        print("\n" + "=" * 80)
        print("ANALYSE TERMINÉE")
        print("=" * 80)
        
        return self.results


class ResultsExporter:
    """Export des résultats vers Excel et graphiques"""
    
    def __init__(self, analyzer, output_dir):
        self.analyzer = analyzer
        self.output_dir = Path(output_dir)
        
    def create_summary_report(self):
        """Création du rapport de synthèse Excel"""
        wb = Workbook()
        
        # Feuille 1: Résumé des classifications
        ws_summary = wb.active
        ws_summary.title = "Résumé Classifications"
        
        headers = ["Indicateur", "Classification", "Tendance", "Saisonnalité", 
                   "Volatilité", "Modèle", "R²", "Précision"]
        ws_summary.append(headers)
        
        # Style des headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for indicator in self.analyzer.indicator_cols:
            if indicator in self.analyzer.classifications:
                classif = self.analyzer.classifications[indicator]
                model_info = self.analyzer.models.get(indicator, {})
                
                row = [
                    indicator,
                    classif['classification'],
                    classif['trend']['trend_type'],
                    'Oui' if classif['seasonality']['has_seasonality'] else 'Non',
                    classif['volatility']['volatility_type'],
                    model_info.get('name', 'N/A'),
                    model_info.get('metrics', {}).get('r2', 0),
                    model_info.get('metrics', {}).get('accuracy_category', 'N/A')
                ]
                ws_summary.append(row)
        
        # Ajuster les largeurs
        for column in ws_summary.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_summary.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Feuille 2: Métriques détaillées
        ws_metrics = wb.create_sheet("Métriques Modèles")
        ws_metrics.append(["Indicateur", "Modèle", "MAE", "RMSE", "MAPE (%)", "R²", "Catégorie"])
        
        for cell in ws_metrics[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for indicator, model_info in self.analyzer.models.items():
            metrics = model_info['metrics']
            row = [
                indicator,
                model_info['name'],
                metrics.get('mae', 'N/A'),
                metrics.get('rmse', 'N/A'),
                metrics.get('mape', 'N/A'),
                metrics.get('r2', 'N/A'),
                metrics.get('accuracy_category', 'N/A')
            ]
            ws_metrics.append(row)
        
        # Feuille 3: Corrélations
        if self.analyzer.results.get('correlations'):
            ws_corr = wb.create_sheet("Corrélations")
            corr_data = self.analyzer.results['correlations']
            
            if corr_data['strong_correlations']:
                ws_corr.append(["Indicateur 1", "Indicateur 2", "Corrélation", "Type"])
                
                for cell in ws_corr[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for corr in corr_data['strong_correlations']:
                    ws_corr.append([
                        corr['indicator_1'],
                        corr['indicator_2'],
                        corr['correlation'],
                        corr['relationship']
                    ])
        
        # Feuille 4: Projections
        ws_forecast = wb.create_sheet("Projections")
        
        # Créer une table avec toutes les projections
        if self.analyzer.forecasts:
            # Header avec dates
            all_dates = []
            for forecast in self.analyzer.forecasts.values():
                all_dates.extend(forecast.index)
            
            if all_dates:
                all_dates = sorted(set(all_dates))
                header_row = ["Indicateur"] + [str(d) for d in all_dates]
                ws_forecast.append(header_row)
                
                for cell in ws_forecast[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Données de projection
                for indicator, forecast in self.analyzer.forecasts.items():
                    row = [indicator]
                    for date in all_dates:
                        if date in forecast.index:
                            row.append(forecast[date])
                        else:
                            row.append("")
                    ws_forecast.append(row)
        
        # Sauvegarder
        output_file = self.output_dir / "analyse_series_chronologiques.xlsx"
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        print(f"\n✓ Rapport Excel créé: {output_file}")
        
        return output_file
    
    def create_visualization_dashboard(self):
        """Création des visualisations avec Plotly"""
        
        # Graphique 1: Classification des séries
        classifications_count = {}
        for classif in self.analyzer.classifications.values():
            cat = classif['classification']
            classifications_count[cat] = classifications_count.get(cat, 0) + 1
        
        fig_classif = go.Figure(data=[
            go.Bar(x=list(classifications_count.keys()), 
                   y=list(classifications_count.values()),
                   marker_color='#1F4E78')
        ])
        fig_classif.update_layout(
            title="Distribution des classifications de séries",
            xaxis_title="Type de comportement",
            yaxis_title="Nombre d'indicateurs",
            template="plotly_white"
        )
        fig_classif.write_html(self.output_dir / "classification_distribution.html")
        
        # Graphique 2: Précision des modèles
        model_accuracy = []
        indicators_list = []
        
        for indicator, model_info in self.analyzer.models.items():
            r2 = model_info['metrics'].get('r2', 0)
            if not np.isnan(r2):
                model_accuracy.append(r2)
                indicators_list.append(indicator)
        
        if model_accuracy:
            fig_accuracy = go.Figure(data=[
                go.Bar(x=indicators_list, y=model_accuracy,
                       marker_color='#2E75B5',
                       text=[f"{val:.2f}" for val in model_accuracy],
                       textposition='auto')
            ])
            fig_accuracy.update_layout(
                title="Précision des modèles (R²)",
                xaxis_title="Indicateur",
                yaxis_title="R² Score",
                template="plotly_white",
                yaxis=dict(range=[0, 1])
            )
            fig_accuracy.write_html(self.output_dir / "model_accuracy.html")
        
        # Graphique 3: Matrice de corrélation
        if self.analyzer.results.get('correlations'):
            corr_matrix = self.analyzer.results['correlations']['correlation_matrix']
            
            fig_corr = go.Figure(data=go.Heatmap(
                z=corr_matrix.values,
                x=corr_matrix.columns,
                y=corr_matrix.columns,
                colorscale='RdBu',
                zmid=0,
                text=corr_matrix.values.round(2),
                texttemplate='%{text}',
                textfont={"size": 10}
            ))
            fig_corr.update_layout(
                title="Matrice de corrélation entre indicateurs",
                template="plotly_white",
                width=800,
                height=800
            )
            fig_corr.write_html(self.output_dir / "correlation_matrix.html")
        
        print(f"✓ Visualisations créées dans: {self.output_dir}")
    
    def create_individual_charts(self):
        """Création des graphiques individuels pour chaque série"""
        
        for indicator in self.analyzer.indicator_cols:
            if indicator not in self.analyzer.classifications:
                continue
            
            series = self.analyzer.df[indicator]
            
            fig = make_subplots(
                rows=2, cols=1,
                subplot_titles=(f'{indicator} - Données historiques et projection',
                              'Analyse des résidus'),
                vertical_spacing=0.15,
                row_heights=[0.7, 0.3]
            )
            
            # Série historique
            fig.add_trace(
                go.Scatter(x=series.index, y=series.values,
                          mode='lines+markers',
                          name='Historique',
                          line=dict(color='#1F4E78', width=2)),
                row=1, col=1
            )
            
            # Projection si disponible
            if indicator in self.analyzer.forecasts:
                forecast = self.analyzer.forecasts[indicator]
                fig.add_trace(
                    go.Scatter(x=forecast.index, y=forecast.values,
                              mode='lines+markers',
                              name='Projection',
                              line=dict(color='#FF6B6B', width=2, dash='dash')),
                    row=1, col=1
                )
            
            # Résidus du modèle
            if indicator in self.analyzer.models:
                model_info = self.analyzer.models[indicator]
                # Calcul simple des résidus pour visualisation
                try:
                    fitted_values = series.rolling(window=3).mean()
                    residuals = series - fitted_values
                    
                    fig.add_trace(
                        go.Scatter(x=residuals.index, y=residuals.values,
                                  mode='markers',
                                  name='Résidus',
                                  marker=dict(color='#95a5a6', size=5)),
                        row=2, col=1
                    )
                    fig.add_hline(y=0, line_dash="dash", line_color="gray", row=2, col=1)
                except:
                    pass
            
            # Mise en forme
            classification = self.analyzer.classifications[indicator]['classification']
            model_name = self.analyzer.models.get(indicator, {}).get('name', 'N/A')
            r2_score = self.analyzer.models.get(indicator, {}).get('metrics', {}).get('r2', 0)
            
            fig.update_layout(
                title=f"{indicator}<br><sub>Classification: {classification} | Modèle: {model_name} | R²: {r2_score:.3f}</sub>",
                template="plotly_white",
                height=800,
                showlegend=True
            )
            
            fig.update_xaxes(title_text="Date", row=2, col=1)
            fig.update_yaxes(title_text="Valeur", row=1, col=1)
            fig.update_yaxes(title_text="Résidus", row=2, col=1)
            
            # Sauvegarder
            safe_filename = indicator.replace('/', '_').replace(' ', '_')
            fig.write_html(self.output_dir / f"serie_{safe_filename}.html")
        
        print(f"✓ Graphiques individuels créés pour {len(self.analyzer.indicator_cols)} indicateurs")


def load_excel_data(input_dir):
    """Chargement des fichiers Excel du répertoire d'entrée"""
    input_path = Path(input_dir)
    excel_files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))
    
    if not excel_files:
        raise FileNotFoundError(f"Aucun fichier Excel trouvé dans {input_dir}")
    
    print(f"\n {len(excel_files)} fichier(s) Excel trouvé(s)")
    
    # Charger le premier fichier (ou combiner si plusieurs)
    df = pd.read_excel(excel_files[0])
    print(f" - {excel_files[0].name}: {df.shape[0]} lignes, {df.shape[1]} colonnes")
    
    return df


def main():
    """Fonction principale"""
    print("\n" + "="*80)
    print("SYSTÈME D'ANALYSE DE SÉRIES CHRONOLOGIQUES P&L")
    print("="*80)
    
    try:
        # Chargement des données
        df = load_excel_data(INPUT_DIR)
        
        # Identifier la colonne de date (première colonne ou colonne contenant 'date')
        date_col = None
        for col in df.columns:
            if 'date' in col.lower() or df[col].dtype == 'datetime64[ns]':
                date_col = col
                break
        
        if date_col is None and df.columns[0] not in df.select_dtypes(include=[np.number]).columns:
            date_col = df.columns[0]
        
        print(f"\n Colonne de date identifiée: {date_col}")
        
        # Colonnes d'indicateurs (toutes sauf date)
        indicator_cols = [col for col in df.columns if col != date_col]
        indicator_cols = [col for col in indicator_cols if df[col].dtype in [np.float64, np.int64, float, int]]
        
        print(f"{len(indicator_cols)} indicateurs à analyser")
        
        # Analyse
        analyzer = TimeSeriesAnalyzer(df, date_col=date_col, indicator_cols=indicator_cols)
        analyzer.analyze_all()
        
        # Export des résultats
        exporter = ResultsExporter(analyzer, OUTPUT_DIR)
        exporter.create_summary_report()
        exporter.create_visualization_dashboard()
        exporter.create_individual_charts()
        
        print("\n" + "="*80)
        print("✓ ANALYSE TERMINÉE AVEC SUCCÈS")
        print(f"Tous les résultats sont disponibles dans: {OUTPUT_DIR}")
        print("="*80 + "\n")
        
    except Exception as e:
        print(f"\n ERREUR: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
